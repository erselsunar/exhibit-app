from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from datetime import date, datetime
from typing import Any

router = APIRouter()

COLUMNS = ["client", "city", "venue", "expo name", "expo start date", "expo end date", "expo management", "booth no"]


def _parse_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        d = value.date() if isinstance(value, datetime) else value
        return d.strftime("%d.%m.%Y")
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d.%m.%Y")
        except ValueError:
            pass
    return s


@router.post("/bulk-upload/preview")
async def bulk_upload_preview(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Sadece .xlsx veya .xls dosyası yükleyebilirsiniz.")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Dosya boş.")

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    col_map: dict[str, int] = {}
    missing_cols = []
    for col in COLUMNS:
        try:
            col_map[col] = headers.index(col)
        except ValueError:
            missing_cols.append(col)

    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail=f"Sütun bulunamadı: {', '.join(missing_cols)}. Beklenen: {', '.join(COLUMNS)}",
        )

    result = []
    for row in rows[1:]:
        def cell(col: str) -> str:
            v = row[col_map[col]]
            return str(v).strip() if v is not None else ""

        record = {
            "client": cell("client"),
            "city": cell("city"),
            "venue": cell("venue"),
            "expo_name": cell("expo name"),
            "expo_start_date": _parse_date(row[col_map["expo start date"]]),
            "expo_end_date": _parse_date(row[col_map["expo end date"]]),
            "expo_management": cell("expo management"),
            "booth_no": cell("booth no"),
        }

        if not any(record.values()):
            continue

        result.append(record)

    return {"rows": result, "total": len(result)}


@router.get("/bulk-upload/template")
def bulk_upload_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    headers = ["Client", "City", "Venue", "Expo Name", "Expo Start Date", "Expo End Date", "Expo Management", "Booth No"]

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 16)

    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exhibit-template.xlsx"},
    )
